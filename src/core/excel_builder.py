import re
import copy
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.models import CodePrixBPU, SectionBPU, LigneConsolidee

# Motif generique acceptant aussi bien les codes "plats" (A1, J7, M8, Y3.1)
# que les codes hierarchiques a tiret (A1-1, C2-1, G3-4, AA1-2...). Le groupe 1
# capture le prefixe alphabetique (1 ou 2 lettres) qui servira de cle de
# regroupement de section, meme quand aucune ligne d'en-tete de section n'est
# detectee dans le fichier source.
LEAF_PATTERN = re.compile(r"^([A-Z]{1,2})(\d+(?:\.\d+)?)(?:-(\d+(?:\.\d+)?))?$")
SECTION_TITLE_PATTERN = re.compile(r"^([A-Z]{1,2})$")

KEYWORDS = {
    "numero": [
        "n\u00b0 de prix", "n de prix", "numero de prix", "code prix",
        "n\u00b0prix", "n\u00b0 prix", "n prix",
    ],
    "designation": ["designation", "d\u00e9signation", "libelle", "libell\u00e9"],
    "unite": ["unite", "unit\u00e9", "unites", "unit\u00e9s"],
    "pu": ["prix unitaire", "p.u.", "pu ht", "pu (ht)"],
}
EXCLUDE_PU = ["designation", "d\u00e9signation", "en toutes lettres"]

HORS_BPU_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")


def _first_line(value: str) -> str:
    return (value or "").split("\n")[0].strip().lower()


def detect_columns(ws: Worksheet, header_row: int = 1, max_col: int = 30) -> dict:
    detected = {"numero": None, "designation": None, "unite": None, "pu": None}
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        raw = str(cell.value or "")
        first_line = _first_line(raw)
        if not first_line:
            continue
        for key, kws in KEYWORDS.items():
            if key == "pu" and any(ex in first_line for ex in EXCLUDE_PU):
                continue
            if detected[key] is None and any(kw in first_line for kw in kws):
                detected[key] = col
    warnings = [k for k, v in detected.items() if v is None]
    if warnings:
        detected["_defaults_used"] = True
        detected["numero"] = detected["numero"] or 1
        detected["designation"] = detected["designation"] or 2
        detected["unite"] = detected["unite"] or 3
        detected["pu"] = detected["pu"] or 4
    else:
        detected["_defaults_used"] = False
    detected["_missing"] = warnings
    return detected


def parse_bpu_structure(path: Path, sheet_name: str | None = None) -> tuple[dict[str, CodePrixBPU], list[SectionBPU], dict]:
    """Parseur agnostique a la profondeur de numerotation du BPUF.

    Passe 1 : on releve tous les codes candidats matchant LEAF_PATTERN, qu'ils
    soient plats (A1) ou a tiret (A1-1).
    Passe 2 : un code plat (ex "A1") est ecarte des feuilles s'il existe au
    moins un code a tiret partageant le meme prefixe plat (ex "A1-1", "A1-2")
    -> c'est alors une ligne de sous-categorie/en-tete, pas un prix reel.
    Un code plat SANS aucun enfant a tiret (ex "T1", "R1" dans certains BPUF)
    reste una feuille valide.
    Les sections sont ensuite reconstituees en regroupant les feuilles
    consecutives partageant le meme prefixe de lettres, independamment de la
    presence ou non d'une ligne d'en-tete explicite dans le fichier source.
    """
    wb = load_workbook(str(path), data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    cols = detect_columns(ws)

    candidates = []
    section_titles: dict[str, str] = {}
    max_row = ws.max_row

    for row in range(2, max_row + 1):
        raw_num = str(ws.cell(row=row, column=cols["numero"]).value or "").strip()
        if not raw_num:
            continue

        title_match = SECTION_TITLE_PATTERN.match(raw_num)
        if title_match:
            letters = title_match.group(1)
            title_text = str(ws.cell(row=row, column=cols["designation"]).value or "").strip()
            if letters not in section_titles and title_text:
                section_titles[letters] = title_text
            continue

        m = LEAF_PATTERN.match(raw_num)
        if not m:
            continue
        letters, num1, num2 = m.group(1), m.group(2), m.group(3)
        bare_text = f"{letters}{num1}"

        designation_raw = str(ws.cell(row=row, column=cols["designation"]).value or "")
        unite = str(ws.cell(row=row, column=cols["unite"]).value or "").strip()
        pu_val = ws.cell(row=row, column=cols["pu"]).value
        try:
            pu_ref = float(pu_val) if pu_val is not None else 0.0
        except (TypeError, ValueError):
            pu_ref = 0.0

        candidates.append({
            "row": row, "code": raw_num, "bare": bare_text, "letters": letters,
            "has_hyphen": num2 is not None, "designation": designation_raw,
            "unite": unite, "pu": pu_ref,
        })

    bare_with_children = {c["bare"] for c in candidates if c["has_hyphen"]}
    leaves = [c for c in candidates if c["has_hyphen"] or c["bare"] not in bare_with_children]

    codes: dict[str, CodePrixBPU] = {}
    for c in leaves:
        intitule_court = (
            _first_line(c["designation"]).split("pu :")[0].split("l'heure")[0].strip()
            or c["designation"].strip()
        )
        codes[c["code"]] = CodePrixBPU(
            code=c["code"], intitule_court=intitule_court, unite=c["unite"],
            pu_reference=c["pu"], section=c["letters"], ligne_source=c["row"],
        )

    sections: list[SectionBPU] = []
    if leaves:
        current_letters = leaves[0]["letters"]
        section_first_row = leaves[0]["row"]
        section_last_row = leaves[0]["row"]
        for c in leaves[1:]:
            if c["letters"] == current_letters:
                section_last_row = c["row"]
                continue
            sections.append(SectionBPU(
                lettre=current_letters, ligne_entete=section_first_row,
                ligne_debut=section_first_row, ligne_fin=section_last_row,
            ))
            current_letters = c["letters"]
            section_first_row = c["row"]
            section_last_row = c["row"]
        sections.append(SectionBPU(
            lettre=current_letters, ligne_entete=section_first_row,
            ligne_debut=section_first_row, ligne_fin=section_last_row,
        ))

    meta = {
        "cols": cols, "nb_sections": len(sections), "nb_codes": len(codes),
        "section_titles": section_titles,
    }
    wb.close()
    return codes, sections, meta


def _copy_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = copy.copy(src_cell.number_format)
        dst_cell.alignment = copy.copy(src_cell.alignment)


def generate_dqe_from_bpu(
    bpu_path: Path,
    consolidation: dict[str, LigneConsolidee],
    output_path: Path,
    qte_defaut_absent: float = 0.0,
    omit_empty_sections: bool = False,
    sheet_name: str | None = None,
) -> dict:
    codes_bpu, sections, meta = parse_bpu_structure(bpu_path, sheet_name)
    section_titles = meta.get("section_titles", {})
    cols = meta["cols"]

    dqe_wb = Workbook()
    dqe_ws = dqe_wb.active
    dqe_ws.title = "DQE"

    headers = ["N\u00b0 Prix", "Intitul\u00e9 court", "Unit\u00e9", "Prix Unitaire", "Quantit\u00e9", "Total"]
    for c, h in enumerate(headers, start=1):
        dqe_ws.cell(row=1, column=c, value=h)

    current_row = 2
    section_totals: list[tuple[str, int]] = []
    not_found_in_history = []

    for section in sections:
        codes_section = [c for c, v in codes_bpu.items() if v.section == section.lettre]
        if not codes_section:
            continue
        if omit_empty_sections and not any(c in consolidation for c in codes_section):
            continue

        titre = section_titles.get(section.lettre, f"SECTION {section.lettre}")
        header_cell = dqe_ws.cell(row=current_row, column=1, value=titre)
        header_cell.font = header_cell.font.copy(bold=True)
        current_row += 1
        section_start = current_row

        for code in sorted(codes_section):
            bpu_info = codes_bpu[code]
            consolide = consolidation.get(code)
            statut_constate = consolide is not None
            if not statut_constate:
                not_found_in_history.append(code)

            pu = consolide.pu_dqe if consolide else bpu_info.pu_reference
            qte = consolide.qte_dqe if consolide else qte_defaut_absent

            dqe_ws.cell(row=current_row, column=1, value=code)
            dqe_ws.cell(row=current_row, column=2, value=bpu_info.intitule_court)
            dqe_ws.cell(row=current_row, column=3, value=bpu_info.unite)
            pu_cell = dqe_ws.cell(row=current_row, column=4, value=round(pu, 4))
            dqe_ws.cell(row=current_row, column=5, value=round(qte, 3))
            total_cell = dqe_ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}")

            pu_cell.number_format = "#,##0.00 \u20ac"
            total_cell.number_format = "#,##0.00 \u20ac"

            if not statut_constate:
                for col_idx in range(1, 7):
                    cell = dqe_ws.cell(row=current_row, column=col_idx)
                    cell.font = cell.font.copy(italic=True, color="808080")

            current_row += 1

        section_end = current_row - 1
        if section_end >= section_start:
            dqe_ws.cell(row=current_row, column=2, value=f"TOTAL SECTION {section.lettre}")
            total_section_cell = dqe_ws.cell(row=current_row, column=6, value=f"=SUM(F{section_start}:F{section_end})")
            total_section_cell.number_format = "#,##0.00 \u20ac"
            section_totals.append((section.lettre, current_row))
            current_row += 1
        current_row += 1

    codes_hors_bpu = sorted(c for c in consolidation if c not in codes_bpu)
    if codes_hors_bpu:
        dqe_ws.cell(row=current_row, column=1, value="HORS BPUF - A VERIFIER / CORRIGER MANUELLEMENT")
        for col_idx in range(1, 7):
            dqe_ws.cell(row=current_row, column=col_idx).fill = HORS_BPU_FILL
        current_row += 1
        hors_bpu_start = current_row

        for code in codes_hors_bpu:
            consolide = consolidation[code]
            dqe_ws.cell(row=current_row, column=1, value=code)
            dqe_ws.cell(row=current_row, column=2, value=consolide.designation)
            dqe_ws.cell(row=current_row, column=3, value=consolide.unite)
            pu_cell = dqe_ws.cell(row=current_row, column=4, value=round(consolide.pu_dqe, 4))
            dqe_ws.cell(row=current_row, column=5, value=round(consolide.qte_dqe, 3))
            total_cell = dqe_ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}")
            pu_cell.number_format = "#,##0.00 \u20ac"
            total_cell.number_format = "#,##0.00 \u20ac"
            for col_idx in range(1, 7):
                dqe_ws.cell(row=current_row, column=col_idx).fill = HORS_BPU_FILL
            current_row += 1

        hors_bpu_end = current_row - 1
        dqe_ws.cell(row=current_row, column=2, value="TOTAL HORS BPUF")
        total_hors_bpu_cell = dqe_ws.cell(row=current_row, column=6, value=f"=SUM(F{hors_bpu_start}:F{hors_bpu_end})")
        total_hors_bpu_cell.number_format = "#,##0.00 \u20ac"
        for col_idx in range(1, 7):
            dqe_ws.cell(row=current_row, column=col_idx).fill = HORS_BPU_FILL
        section_totals.append(("HORS_BPU", current_row))
        current_row += 2

    if section_totals:
        formula = "=" + "+".join(f"F{row}" for _, row in section_totals)
        dqe_ws.cell(row=current_row, column=2, value="TOTAL GENERAL HT")
        total_general_cell = dqe_ws.cell(row=current_row, column=6, value=formula)
        total_general_cell.number_format = "#,##0.00 \u20ac"

    for col_idx, width in enumerate([12, 55, 10, 15, 12, 15], start=1):
        dqe_ws.column_dimensions[get_column_letter(col_idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    dqe_wb.save(str(output_path))

    return {
        "nb_sections": len(sections),
        "nb_codes_bpu": len(codes_bpu),
        "nb_codes_constates": len(consolidation),
        "codes_absents_historique": not_found_in_history,
        "codes_hors_bpu": codes_hors_bpu,
        "colonnes_detectees": cols,
        "output_path": str(output_path),
    }


def verify_formulas(output_path: Path) -> list[str]:
    wb = load_workbook(str(output_path), data_only=False)
    ws = wb.active
    anomalies = []
    for row in range(2, ws.max_row + 1):
        d_val = ws.cell(row=row, column=4).value
        e_val = ws.cell(row=row, column=5).value
        f_val = ws.cell(row=row, column=6).value
        if isinstance(d_val, (int, float)) and isinstance(e_val, (int, float)):
            if not (isinstance(f_val, str) and f_val.startswith("=")):
                anomalies.append(f"Ligne {row}: formule totale manquante")
    wb.close()
    return anomalies
