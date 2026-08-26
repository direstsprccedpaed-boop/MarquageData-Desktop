import re
import csv
import json
import hashlib
from pathlib import Path
from typing import Optional

import pdfplumber
import docx
from openpyxl import load_workbook

from core.models import CommandeExtraite, LigneCommande, DocumentIngere
from core.paths import app_data_dir

TEMPLATES_PATH = app_data_dir() / "data" / "extraction_templates.json"

LINE_KEYWORDS = {
    "code": ["code prix", "n\u00b0 prix", "n prix", "n\u00b0 de prix", "reference", "r\u00e9f\u00e9rence", "r\u00e9f.", "article", "code"],
    "designation": ["designation", "d\u00e9signation", "libelle", "libell\u00e9", "intitul\u00e9", "d\u00e9nomination"],
    "quantite": ["quantite", "quantit\u00e9", "qte", "qt\u00e9", "qty"],
    "unite": ["unite", "unit\u00e9", "un.", "u.", "unit"],
    "pu": ["prix unitaire", "p.u.", "pu ht", "prix u.", "prix unit."],
}

DATE_PATTERN = re.compile(r"(\d{2}[/\-.]\d{2}[/\-.]\d{4}|\d{4}-\d{2}-\d{2})")
NUM_COMMANDE_PATTERN = re.compile(r"(?:n\u00b0|num[e\u00e9]ro)\s*(?:de\s*)?commande\s*[:\-]?\s*([A-Za-z0-9\-\/]+)", re.IGNORECASE)
SECTEUR_PATTERN = re.compile(r"(CEI\s+[A-Z\u00c9\u00c8\u00c0][\w\-\s]{2,30}|SREX[\-\s]?\w*)", re.IGNORECASE)
NUM_VALUE_PATTERN = re.compile(r"-?\d+(?:[.,]\d+)?")


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def fingerprint_header(header: list[str]) -> str:
    normalized = "|".join(_normalize(h) for h in header)
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:16]


def load_templates() -> dict:
    if not TEMPLATES_PATH.exists():
        return {}
    return json.loads(TEMPLATES_PATH.read_text(encoding="utf-8"))


def save_templates(templates: dict) -> None:
    TEMPLATES_PATH.parent.mkdir(parents=True, exist_ok=True)
    TEMPLATES_PATH.write_text(json.dumps(templates, ensure_ascii=False, indent=2), encoding="utf-8")


def confirm_template(fingerprint: str, header: list[str], mapping: dict) -> None:
    templates = load_templates()
    templates[fingerprint] = {"header": header, "mapping": mapping}
    save_templates(templates)


def list_templates() -> dict:
    return load_templates()


def delete_template(fingerprint: str) -> None:
    templates = load_templates()
    templates.pop(fingerprint, None)
    save_templates(templates)


def detect_line_columns(header: list[str]) -> dict:
    detected = {"code": None, "designation": None, "quantite": None, "unite": None, "pu": None}
    for idx, cell in enumerate(header):
        norm = _normalize(cell)
        if not norm:
            continue
        for key, kws in LINE_KEYWORDS.items():
            if detected[key] is None and any(kw in norm for kw in kws):
                detected[key] = idx
    return detected


def _parse_number(raw) -> Optional[float]:
    if raw is None:
        return None
    text = str(raw).strip().replace("\u00a0", "").replace(" ", "").replace("\u20ac", "")
    if not text:
        return None
    text = text.replace(",", ".")
    m = NUM_VALUE_PATTERN.search(text)
    if not m:
        return None
    try:
        return float(m.group())
    except ValueError:
        return None


def extract_tables_xlsx(path: Path) -> list[list[list[str]]]:
    wb = load_workbook(str(path), data_only=True, read_only=True)
    tables = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                rows.append(cells)
        if rows:
            tables.append(rows)
    wb.close()
    return tables


def extract_tables_csv(path: Path) -> list[list[list[str]]]:
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
                except csv.Error:
                    dialect = csv.excel
                rows = [row for row in csv.reader(f, dialect) if any(c.strip() for c in row)]
            return [rows] if rows else []
        except (UnicodeDecodeError, LookupError):
            continue
    return []


def extract_tables_docx(path: Path) -> list[list[list[str]]]:
    document = docx.Document(str(path))
    tables = []
    for table in document.tables:
        rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
        rows = [r for r in rows if any(c.strip() for c in r)]
        if rows:
            tables.append(rows)
    return tables


def extract_table_by_words_pdf(path: Path, y_tolerance: float = 3.0, x_tolerance: float = 14.0) -> list[list[list[str]]]:
    """Reconstruction de tableau par position des mots (coordonnees x0/top),
    independamment de toute detection de grille/bordures.

    Necessaire car de nombreux bons de commande administratifs positionnent
    le texte en blocs (colonnes visuelles) sans veritables traits de tableau :
    pdfplumber.extract_tables() lit alors le contenu dans un ordre plus ou
    moins colonne-par-colonne au lieu de ligne-par-ligne, ce qui desynchronise
    quantites/prix par rapport aux codes/designations. Confirme sur un cas
    reel (bon de commande DIR Est, page 2 : la table N\u00b0 de prix/Designation/
    Unite/Qtes/Prix unitaire sortait de extract_tables() avec les quantites et
    prix totaux completement decorreles de leur ligne d'origine).

    Principe : regrouper les mots en lignes par proximite verticale, deduire
    les positions de colonnes ("slots") a partir de la distribution globale
    des coordonnees x0 de la page, puis reaffecter chaque mot a la colonne la
    plus proche en concatenant les mots d'une meme cellule."""
    tables: list[list[list[str]]] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            try:
                words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            except Exception:
                continue
            if not words:
                continue

            words_sorted = sorted(words, key=lambda w: (w["top"], w["x0"]))
            rows: list[list[dict]] = []
            current_row: list[dict] = []
            current_top: float | None = None
            for w in words_sorted:
                if current_top is None or abs(w["top"] - current_top) <= y_tolerance:
                    current_row.append(w)
                    current_top = w["top"] if current_top is None else current_top
                else:
                    rows.append(current_row)
                    current_row = [w]
                    current_top = w["top"]
            if current_row:
                rows.append(current_row)

            all_x0 = sorted({round(w["x0"], 1) for w in words})
            column_slots: list[float] = []
            for x in all_x0:
                if not column_slots or x - column_slots[-1] > x_tolerance:
                    column_slots.append(x)

            grid: list[list[str]] = []
            for row_words in rows:
                row_words_sorted = sorted(row_words, key=lambda w: w["x0"])
                cells = ["" for _ in column_slots]
                for w in row_words_sorted:
                    slot_idx = min(range(len(column_slots)), key=lambda i: abs(column_slots[i] - w["x0"]))
                    cells[slot_idx] = (cells[slot_idx] + " " + w["text"]).strip()
                if any(c.strip() for c in cells):
                    grid.append(cells)

            if grid:
                tables.append(grid)
    return tables


def extract_tables_pdf(path: Path) -> list[list[list[str]]]:
    """Combine deux strategies d'extraction PDF concurrentes : la detection
    de grille classique de pdfplumber (fiable quand le PDF a de vraies
    bordures de cellules), et une reconstruction par position de mots
    (fiable quand le tableau est fait de blocs de texte positionnes, sans
    bordures detectables). Les candidats des deux methodes sont ensuite
    departages par le systeme de score commun (select_best_table) - aucune
    des deux methodes n'est privilegiee a priori."""
    tables: list[list[list[str]]] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                for raw_table in page.extract_tables() or []:
                    rows = [["" if c is None else str(c).strip() for c in row] for row in raw_table]
                    rows = [r for r in rows if any(c.strip() for c in r)]
                    if rows:
                        tables.append(rows)
    except Exception:
        pass

    try:
        tables.extend(extract_table_by_words_pdf(path))
    except Exception:
        pass

    return tables


EXTRACTORS = {
    ".xlsx": extract_tables_xlsx,
    ".csv": extract_tables_csv,
    ".docx": extract_tables_docx,
    ".pdf": extract_tables_pdf,
}


def extract_metadata(texte: str) -> dict:
    date_commande = ""
    date_match = DATE_PATTERN.search(texte)
    if date_match:
        raw = date_match.group(1).replace("/", "-").replace(".", "-")
        parts = raw.split("-")
        if len(parts[0]) == 4:
            date_commande = raw
        elif len(parts) == 3 and len(parts[2]) == 4:
            date_commande = f"{parts[2]}-{parts[1]}-{parts[0]}"

    num_match = NUM_COMMANDE_PATTERN.search(texte)
    numero_commande = num_match.group(1) if num_match else ""

    secteur_match = SECTEUR_PATTERN.search(texte)
    secteur = secteur_match.group(1).strip() if secteur_match else ""

    return {"numero_commande": numero_commande, "date_commande": date_commande, "secteur": secteur}


def _keyword_score(row: list[str]) -> tuple[int, dict]:
    mapping = detect_line_columns(row)
    score = sum(1 for v in mapping.values() if v is not None)
    return score, mapping


def _numeric_evidence(table: list[list[str]], header_idx: int, max_check_rows: int = 10) -> int:
    data_rows = table[header_idx + 1: header_idx + 1 + max_check_rows]
    if not data_rows:
        return 0
    nb_cols = max((len(r) for r in data_rows), default=0)
    numeric_cols = 0
    for col in range(nb_cols):
        values = [r[col] for r in data_rows if col < len(r) and r[col].strip()]
        if not values:
            continue
        numeric_count = sum(1 for v in values if _parse_number(v) is not None)
        if numeric_count >= max(2, (len(values) + 1) // 2):
            numeric_cols += 1
    return numeric_cols


def find_header_row(table: list[list[str]], max_scan: int = 4) -> tuple[int, int, dict]:
    best_idx, best_score, best_mapping = 0, -1, {}
    for idx in range(min(max_scan, len(table))):
        score, mapping = _keyword_score(table[idx])
        if score > best_score:
            best_idx, best_score, best_mapping = idx, score, mapping
    return best_idx, best_score, best_mapping


def score_candidate_table(table: list[list[str]]) -> dict:
    header_idx, kw_score, mapping = find_header_row(table)
    numeric_cols = _numeric_evidence(table, header_idx)
    nb_data_rows = len(table) - header_idx - 1
    total_cells = sum(len(row) for row in table) or 1
    non_empty_cells = sum(1 for row in table for c in row if c.strip())
    non_empty_ratio = non_empty_cells / total_cells

    total_score = (kw_score * 3) + (numeric_cols * 2) + (1 if nb_data_rows >= 2 else -3)
    if non_empty_ratio < 0.15:
        total_score -= 5

    return {
        "header_idx": header_idx, "kw_score": kw_score, "mapping": mapping,
        "numeric_cols": numeric_cols, "nb_data_rows": nb_data_rows,
        "non_empty_ratio": round(non_empty_ratio, 2), "total_score": total_score,
    }


def select_best_table(tables: list[list[list[str]]]) -> tuple[Optional[list[list[str]]], Optional[dict]]:
    if not tables:
        return None, None
    scored = [(t, score_candidate_table(t)) for t in tables]
    scored.sort(key=lambda x: x[1]["total_score"], reverse=True)
    best_table, best_info = scored[0]
    if best_info["kw_score"] < 1 and best_info["numeric_cols"] < 2:
        return None, best_info
    return best_table, best_info


def extract_lines_from_table(table: list[list[str]], mapping: dict) -> tuple[list[LigneCommande], int]:
    lignes: list[LigneCommande] = []
    nb_ignorees = 0
    for row in table[1:]:
        code_idx = mapping.get("code")
        if code_idx is None or code_idx >= len(row):
            continue
        code = row[code_idx].strip()
        if not code:
            continue

        desi_idx = mapping.get("designation")
        designation = row[desi_idx].strip() if desi_idx is not None and desi_idx < len(row) else ""

        qte_idx = mapping.get("quantite")
        qte = _parse_number(row[qte_idx]) if qte_idx is not None and qte_idx < len(row) else None

        unite_idx = mapping.get("unite")
        unite = row[unite_idx].strip() if unite_idx is not None and unite_idx < len(row) else ""

        pu_idx = mapping.get("pu")
        pu = _parse_number(row[pu_idx]) if pu_idx is not None and pu_idx < len(row) else None

        if qte is None or pu is None:
            nb_ignorees += 1
            continue
        try:
            lignes.append(LigneCommande(
                code_prix_source=code, designation=designation, quantite=qte, unite=unite, pu_ht=pu,
            ))
        except Exception:
            nb_ignorees += 1
    return lignes, nb_ignorees


def extract_document_local(doc: DocumentIngere, templates: dict | None = None) -> tuple[CommandeExtraite | None, dict]:
    templates = templates if templates is not None else load_templates()
    path = Path(doc.chemin)
    ext = path.suffix.lower()
    extractor = EXTRACTORS.get(ext)
    if not extractor:
        return None, {"erreur": f"format non supporte pour extraction locale: {ext}"}

    try:
        tables = extractor(path)
    except Exception as exc:
        return None, {"erreur": str(exc)}

    if not tables:
        return None, {"erreur": "aucun tableau detecte dans ce document \u2014 essayez le flux LLM (onglet 2)"}

    table, info = select_best_table(tables)
    if table is None:
        return None, {
            "erreur": "aucune table fiable identifi\u00e9e (probablement une page de garde/r\u00e9capitulatif "
                      "sans donn\u00e9es de prix exploitables) \u2014 utilisez le flux LLM (onglet 2) pour ce document",
        }

    header_idx = info["header_idx"]
    working_table = table[header_idx:]
    header = working_table[0]
    fp = fingerprint_header(header)

    if fp in templates:
        mapping = templates[fp]["mapping"]
        mapping_a_valider = None
    else:
        mapping = info["mapping"]
        mapping_a_valider = {
            "fingerprint": fp, "header": header, "mapping": mapping,
            "preview_rows": working_table[1:4],
        }

    essentiels_ok = mapping.get("code") is not None and mapping.get("quantite") is not None and mapping.get("pu") is not None
    lignes, nb_ignorees = extract_lines_from_table(working_table, mapping) if essentiels_ok else ([], 0)

    texte_meta = "\n".join(" ".join(row) for row in table[:3]) + "\n" + doc.texte_brut[:500]
    meta = extract_metadata(texte_meta)

    commande = CommandeExtraite(
        doc_id=doc.doc_id,
        numero_commande=meta["numero_commande"],
        date_commande=meta["date_commande"] or "2000-01-01",
        secteur=meta["secteur"],
        lignes=lignes,
    )

    rapport = {
        "nb_lignes": len(lignes),
        "nb_lignes_ignorees": nb_ignorees,
        "fingerprint": fp,
        "mapping_a_valider": mapping_a_valider,
        "metadata_incomplete": not (meta["numero_commande"] and meta["date_commande"] and meta["secteur"]),
    }
    return commande, rapport


def extract_batch_local(docs: list[DocumentIngere]) -> tuple[list[CommandeExtraite], list[dict], dict]:
    templates = load_templates()
    commandes: list[CommandeExtraite] = []
    rapports: list[dict] = []
    gabarits_en_attente: dict[str, dict] = {}

    for doc in docs:
        commande, rapport = extract_document_local(doc, templates=templates)
        rapport["doc_id"] = doc.doc_id
        rapports.append(rapport)
        if rapport.get("mapping_a_valider"):
            fp = rapport["mapping_a_valider"]["fingerprint"]
            entry = gabarits_en_attente.setdefault(fp, {**rapport["mapping_a_valider"], "doc_ids": []})
            entry["doc_ids"].append(doc.doc_id)
        elif commande and (commande.lignes or not rapport.get("erreur")):
            commandes.append(commande)

    return commandes, rapports, gabarits_en_attente
